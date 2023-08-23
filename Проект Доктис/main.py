import math
import os
import ast
import random

import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
import matplotlib.pyplot as plt
from scipy.signal import savgol_filter
from scipy.interpolate import CubicSpline
import numpy as np
from scipy.interpolate import Akima1DInterpolator


from doctors_result import doctor_result_dict

#Функция проверки наличия ошибок
def find_err(df_all_coords, basal_value):

    mode_err = False   #флаг показывающий, что ошибка началась
    #запускаем циклом перебор индексов и строк датафрейма
    for i, point in df_all_coords.iterrows():
        if i == len(df_all_coords) - 1:
            break
        #смотрим есть ли скачок в 20 между соседними точками, или "ошибка началась" и прямая пересекла базальный уровень
        if (abs(df_all_coords.loc[i, 'y'] - df_all_coords.loc[i + 1, 'y']) > 20) or (mode_err and ((df_all_coords.loc[i, 'y'] - basal_value) * (df_all_coords.loc[i + 1, 'y'] - basal_value) <= 0)):
            if not mode_err: #если ошибка не началась
                start_x = df_all_coords.loc[i, 'x'] #сохранение стартовой точки
                mode_err = True  #ошибка началась
            else:
                new_row = {
                    'type': 'error',
                    'start': start_x
                }
                #во все строки датафрейма, в которых х находится в пределах от стартовой точки до точки i+1, записываем в поле type "error"
                df_all_coords.loc[(df_all_coords['x'] >= start_x) & (df_all_coords['x'] <= df_all_coords.loc[i + 1, 'x']), 'type'] = 'error'

                # df_type.loc[len(df_type)] = new_row

                mode_err = False

def universalType(basal_value, df_all_coords): # функция разметки акцелераций,десцелераций,базального ритма

    start_x = df_all_coords.loc[0, 'x'] #задаем начальную точку х
    position = df_all_coords.loc[0, 'y'] - basal_value #вычисляем начальное положение относительно базального уровня
    mode = False #объявляем флаг, сигнализирующий об акцелерации или десцелерации

    for i, point in df_all_coords.iterrows():
        if point.type == 'error':
            continue
        #Проверка есть ли акцелерация или децелерация
        if abs(basal_value - point.y) >= 15:
            mode = True #если текущая точка отличается от базального уровня на 15, то это ацелерация или десцелерация
        #Определение типа сердечного ритма и запись в датафреймы
        if ((point.y - basal_value) * position <= 0) or (i == len(df_all_coords) - 1): #если прямая пересекла базальный уровень, то режим закончился(переходим к следующему отрезку и сохраняем данные)
            new_row = {
                'start': start_x,
                'duration': point.x - start_x  #продолжительность акцелерации или децелерации
            }

            if (position > 0) and mode and (new_row['duration'] > 15):

                df_all_coords.loc[(df_all_coords['x'] >= start_x) & (df_all_coords['x'] < point.x), 'type'] = 'acsel'

            if (position < 0) and mode and (new_row['duration'] > 15):

                df_all_coords.loc[(df_all_coords['x'] >= start_x) & (df_all_coords['x'] < point.x), 'type'] = 'descel'


            start_x = point.x
            position = point.y - basal_value
            mode = False
            #если участок не акцелерация и не десцелерация, то type остается пустым-базальный ритм

def input_df_type(df_type, df_all_coords, basal_value): #функция наполняющая датафрейм df_type данными
    type_x = df_all_coords.loc[0, 'type']
    start_x = df_all_coords.loc[0, 'x']
    for index, row in df_all_coords.iterrows():  # Перебор строк в датафрейме
        # if index == 0:
        #     type_x = row['type']
        #     start_x = row['x']
        if row['type'] != type_x:
            if type_x == '':
                type_x = 'basal'
            #создаем структуру для записи новой строчки в датафрейме df_type
            new_row = {
                'type': type_x,
                'number': df_type.loc[df_type['type'] == type_x, 'number'].count() + 1,
                'start': start_x,
                'duration': df_all_coords.loc[index - 1, 'x'] - start_x,
            }

            if type_x == 'descel':
                new_row['deep'] = basal_value - df_all_coords.loc[(df_all_coords['x'] >= start_x) & (df_all_coords['x'] <= df_all_coords.loc[index - 1, 'x']), 'y'].min()
            else:
                new_row['deep'] = 0


            df_type.loc[len(df_type)] = new_row   #записываем данные из структуры в датафрейм

            type_x = row['type']  #записываем следующий определяемый тип
            start_x = row['x']  #записываем новую стартовую точку для слд типа

        #записываем последний тип
        if index == len(df_all_coords) - 1:
            if type_x == '':
                type_x = 'basal'
            new_row = {
                'type': type_x,
                'number': df_type.loc[df_type['type'] == type_x, 'number'].count() + 1,
                'start': start_x,
                'duration': df_all_coords.loc[index, 'x'] - start_x
            }
            df_type.loc[len(df_type)] = new_row

def print_and_save_graf(name):
    plt.title(name)
    plt.axhline(y=60, color='black', label='Zero', linewidth=0.5)
    plt.savefig('graf/' + name + '.svg')  # Сохранение файла в папку
    # plt.legend()
    # plt.show()

def print_and_save_graf_period(name):
    name = name + ' period'
    plt.title(name)
    # plt.axhline(y=60, color='black', label='Zero', linewidth=0.5)
    plt.savefig('graf/period/' + name + '.jpg')  # Сохранение файла в папку
    # plt.legend()
    # plt.show()

def score_deceleration(df_type):

    if df_type.loc[df_type['type'] == 'descel'].empty:
        return 2

    # print('desel duration', df_type.loc[df_type['type'] == 'descel', 'duration'].max())
    # print('desel deep', df_type.loc[df_type['type'] == 'descel', 'deep'].max())

    # если максимальная продолжительность десцелераций больше 40 и максимальная глубина больше 35
    if df_type.loc[df_type['type'] == 'descel', 'duration'].max() >= 40 and df_type.loc[df_type['type'] == 'descel', 'deep'].max() >= 35:
        return 0

    return 1

def score_acseleration(df_type, df_all_coords):
    if df_type.loc[df_type['type'] == 'acsel'].empty:
        return 0

    #сохранение количества акцелераций на всем промежутке
    count_acsel = df_type.loc[df_type['type'] == 'acsel', 'number'].max()
    full_time = df_all_coords['x'].max()  #сохранение максимального x

    count_acsel_in_30 = count_acsel * 1800 / full_time  # составляем пропорцию для
                                                        # высчитывания количества акцелераций на промежутке в 1800 секунд

    if count_acsel_in_30 >= 2:
        return 2
    if count_acsel_in_30 >= 1:
        return 1
    return 0

def score_variable_baz_ritm(df_type, df_all_coords): #функция для оценки вариабельности
    i_max_duration_row = df_type.loc[df_type['type'] == 'basal', 'duration'].idxmax() # Строка из df_type с максимальной 'duration' и type='basal':
    max_duration_data = df_type.loc[i_max_duration_row]

    start_x = max_duration_data['start']  # назначаем точку старта по х
    end_x = max_duration_data['start'] + 60  # конец по х
    counter = 0
    sum = 0
    diff = 0

    #начиная со стартового х, в течение каждых 30 секунд проводим исследование вариабельности
    #затем делаем сдвиг начальной точки на 30 секунд и берем минимальное значение
    while end_x <= max_duration_data['start'] + max_duration_data['duration']:
        max_y = df_all_coords.loc[(df_all_coords['x'] >= start_x) & (df_all_coords['x'] <= end_x), 'y'].max()
        min_y = df_all_coords.loc[(df_all_coords['x'] >= start_x) & (df_all_coords['x'] <= end_x), 'y'].min()
        # if diff > max_y - min_y or diff == 0:
        #     diff = max_y - min_y
        diff = max_y - min_y
        sum += diff * diff
        counter += 1
        # print('=> ', counter, diff)

        start_x += 5
        end_x += 5

        if end_x > max_duration_data['start']:
            break

    #выполняется в случае, если наибольший отрезок из датафрейм df_type не превышает 30 секунд
    if counter == 0:
        end_x = max_duration_data['start'] + max_duration_data['duration']
        max_y = df_all_coords.loc[(df_all_coords['x'] >= start_x) & (df_all_coords['x'] <= end_x), 'y'].max()
        min_y = df_all_coords.loc[(df_all_coords['x'] >= start_x) & (df_all_coords['x'] <= end_x), 'y'].min()
        diff = max_y - min_y
        sum = diff * diff
        counter += 1

    sum = math.sqrt(sum / counter)
    # sum = diff

    # start_x = max_duration_data['start']  #  назначаем точку старта по х
    # end_x = max_duration_data['start'] + max_duration_data['duration']  # конец по х
    #
    # max_y = df_all_coords.loc[(df_all_coords['x'] >= start_x) & (df_all_coords['x'] <= end_x), 'y'].max()
    # min_y = df_all_coords.loc[(df_all_coords['x'] >= start_x) & (df_all_coords['x'] <= end_x), 'y'].min()
    # diff = max_y - min_y

    # print('max_y - min_y', sum)

    if 10 <= sum <= 25:
        return 2
    if (5 <= sum <= 9) or (sum > 25):
        return 1
    return 0

def score_osceleration(df_type, df_all_coords, Print, name, basal_value):
    max_duration_row = df_type.loc[df_type['type'] == 'basal', 'duration'].idxmax() # Строка из df_type с максимальной длительностью и type='basal'
    max_duration_data = df_type.loc[max_duration_row]
    #рандомно выбираем точку начала исследования(смотрим осцилляцию)
    random_start_x = random.uniform(max_duration_data.start, max_duration_data.start + max_duration_data.duration - 70)

    #делаем дополнительный график с выбранным базальным ритмом и отмечаем на нем исследуемый участок
#graf 2
    # plt.figure(figsize=(14, 7), dpi=120)  # Задаем размеры создаваемого поля
    # filtered_df = df_all_coords[(df_all_coords['x'] > max_duration_data.start) & (df_all_coords['x'] < max_duration_data.start + max_duration_data.duration)]
    # plt.plot(filtered_df['x'], filtered_df['y'])
    # plt.axhline(y=basal_value, color='black', linestyle='--', label='BV 2', linewidth=1)
    # plt.axvline(x=random_start_x, color='red', label='start', linewidth=1)
    # plt.axvline(x=random_start_x + 60, color='red', label='end', linewidth=1)

    peaks = 0

    # Получение подмножества данных на заданном промежутке (random_start_x + 60)
    subset = df_all_coords[(df_all_coords['x'] >= random_start_x) & (df_all_coords['x'] <= random_start_x + 60)]

    # Подсчет "пиков" в столбце 'y'
    mode_up = False
    for i in range(1, len(subset) - 1):
        if subset.iloc[i]['y'] > subset.iloc[i - 1]['y']: #если график растет, то ищем пик
            mode_up = True
        if subset.iloc[i]['y'] > subset.iloc[i + 1]['y'] and mode_up: # если график убывает, а до этого рос-пик найден:)
            mode_up = False
            peaks += 1

            plt.scatter(subset.iloc[i, 0], subset.iloc[i, 1], color='red')

# graf 2
#     print_and_save_graf_period(name) #сохраняем график

    if Print:  #используется для дебагинга
        print()
        print('x range =', random_start_x, random_start_x + 60) #выбранный участок для исследования(начальная/конечная х)
        print('peacs =', peaks)

    if peaks > 6:
        return 2, max_duration_data.start, max_duration_data.start + max_duration_data.duration
    if 3 <= peaks <= 6:
        return 1, max_duration_data.start, max_duration_data.start + max_duration_data.duration
    return 0, max_duration_data.start, max_duration_data.start + max_duration_data.duration

def score_basal_value(basal_value):  # функция для оценки базального ритма
    # print('basal_value', basal_value, (100 <= basal_value < 120))
    if (120 <= basal_value) and (basal_value <= 160):
        return 2
    if (100 <= basal_value < 120) or (160 < basal_value <= 180):
        return 1
    return 0

def ctg_analyze(df_all_coords):  #main function

    #Создание дополнительного поля в основном датафрейме
    df_all_coords['type'] = ''

    global numberPlot
    global bol6, bol7, bol8
    numberPlot += 1
    score = 0
    ret = score
    graf_list = []  #1-54 55-75 76-81 82-100 61,62,63,64,65,66,67,68,69,70,71,72,73,74,75
    # graf_list = [76, 77, 78, 79, 80, 81, 82, 83, 84, 85, 86, 87, 88, 89, 90, 91, 92, 93, 94, 95, 96, 97, 98, 99, 100]  #1-54 55-75 76-81 82-100 61,62,63,64,65,66,67,68,69,70,71,72,73,74,75


    if (numberPlot in graf_list) or (len(graf_list) == 0):
        print(numberPlot, '/ 100')
        # Создание нового датафрейма с типом, номером графика, временем начала различных оценок сердечного ритма, продолжительностью
        df_type = pd.DataFrame(columns=['type', 'number', 'start', 'duration', 'deep'])

        name1 = "График КТГ №" + str(numberPlot)             #Создаем универсальное имя

        basal_value = df_all_coords['y'].quantile(0.5)  #Вычисление базального уровня путем взятия второго квантиля
        find_err(df_all_coords, basal_value) # вызов функции поиска ошибок на графике
        universalType(basal_value, df_all_coords)
        # повторный вызов функции поиска ошибок на графике, тк функция universalType
        # по принципу своей работы игнорирует ошибки-их нужно проставить заново
        find_err(df_all_coords, basal_value)

        #для более точного вычисления базального уровня берем все неразмеченные участки(не акцелерации,не десцелерации,не ошибки)
        #вычисляем их среднее значение
        basal_value = df_all_coords.loc[df_all_coords['type'] == '', 'y'].mean()  # Вычисление БАЗАЛЬНОГО УРОВНЯ
        #стираем все типы, для повторной, более точной разметки с новым базальным уровнем
        df_all_coords.loc[:, 'type'] = ''

        universalType(basal_value, df_all_coords)  # вызов функции разметки акцелераций,десцелераций,базального ритма, с новым базальным уровнем
        find_err(df_all_coords, basal_value)  #вызов функции поиска ошибок на графике с новым базальным уровнем

        # Формирование df_type( наполняем датафрейм df_type данными)
        input_df_type(df_type, df_all_coords, basal_value)

        #вызов функции для расчета пиков на промежутке в 60 секунд
        osceleration, x1, x2 = score_osceleration(df_type, df_all_coords, False, name1, basal_value)
        print('score_osceleration =', osceleration)

# Отрисовка графика в зависимости от того, к какому типу принадлежит прямая
# graf 1
#         plt.figure(figsize=(14, 7),
#                    dpi=120)  # Задаем размеры создаваемого поля
#         color_map = {'acsel': 'orange', 'basal': 'green', '': 'green', 'descel': 'blue', 'error': 'red'}
#
#         # 2
#         plt.axhline(y=basal_value, color='black', linestyle='--',label='BV 2')  # Отрисовка базального уровня на графике
#         plt.axhline(y=basal_value + 15, color='black', label='up BV 2')
#         plt.axhline(y=basal_value - 15, color='black', label='down BV 2')
#         #3
#         for i, row in df_all_coords.iterrows(): # Вывод цветного графика (способ от точки к точке)
#             if i == 0:
#                 continue
#             plt.plot([df_all_coords.loc[i-1, 'x'], df_all_coords.loc[i, 'x']], [df_all_coords.loc[i-1, 'y'], df_all_coords.loc[i, 'y']], color=color_map[row.type])
# #--------
#         plt.axvline(x=x1, color='red', label='start', linewidth=1)
#         plt.axvline(x=x2, color='red', label='start', linewidth=1)
#         print_and_save_graf(name1)  # Вывод графика


        acseleration = score_acseleration(df_type, df_all_coords)
        # print('score_acseleration =', acseleration)
        deceleration = score_deceleration(df_type)
        # print('score_deceleration =', deceleration)
        s_basal_value = score_basal_value(basal_value)
        # print('score_basal_value =', s_basal_value)
        variable_baz_ritm = score_variable_baz_ritm(df_type, df_all_coords)
        # print('score_variable_baz_ritm =', variable_baz_ritm)
        score = osceleration + acseleration + deceleration + s_basal_value + variable_baz_ritm

        if numberPlot < 76:
            if score > 6:
                bol6 += 1
            if score > 7:
                bol7 += 1
            if score > 8:
                bol8 += 1
        else:
            if score <= 6:
                bol6 += 1
            if score <= 7:
                bol7 += 1
            if score <= 8:
                bol8 += 1


        print('score = ', score, 'bol = ', bol6, bol7, bol8)
        ret = ''
        if score > 6:
            return 'хорошее'
        else:
            return 'плохое'
        # ret = str(score) + ' =>' + ' osceleration = ' + str(osceleration) + '\n' + ' acseleration = ' + str(acseleration) + '\n'' deceleration = ' + str(deceleration)  + '\n' + ' s_basal_value = ' + str(s_basal_value) + 'bas val = ' + str(basal_value) + '\n'' variable_baz_ritm = ' + str(variable_baz_ritm)

        # plt.show() #  Выводим график
    return ret


if __name__ == '__main__':

    global numberPlot
    global bol6, bol7, bol8
    numberPlot, bol6, bol7, bol8 = 0, 0, 0, 0
    directory = 'ctg_files'
    program_result_dict = {}
    start_time = datetime.now()
    #проходим циклом по предоставленным файлам с массивами данных по КТГ
    filename_list = os.listdir(directory)
    filename_list.sort(key=lambda x: int(x[:-4]))
    for filename in filename_list:
        f = os.path.join(directory, filename)
        if os.path.isfile(f):
            file = open(f, 'r')
            graph_list = ast.literal_eval(file.read())
            #преобразуем данные в pandas dataframe для дальнейшей обработки
            #преобразовывать в dataframe необязательно, если имеются другие решения можете реализовать их
            x_coords = [i.get('Key') for i in graph_list]
            y_coords = [i.get('Value') for i in graph_list]
            df_all_coords = pd.DataFrame.from_dict({'x': x_coords,
                                                    'y': y_coords,
                                                    })


            #//////////////////////////////////////////////////////////////////////////////////////////////////////////////
            '''здесь вызывается исполнение функции оценивающей КТГ
            при написании кода рекомендуется использование matplotlib или аналоги для визуализации графика, это поможет
            писать весь код в одной функции необязательно - хорошая читаемость кода приветствуется
            программа в результате должна вернуть строку 'хорошее' или 'плохое'
            '''
            program_result = ctg_analyze(df_all_coords)
            #//////////////////////////////////////////////////////////////////////////////////////////////////////////////


            #записывается результирующий словарь с ключами идентичными словарю doctor_result_dict для дальнейшего сравнения
            program_result_dict[filename] = program_result

    #считаем среднее время выполнения оценки одного КТГ
    average_time = (datetime.now() - start_time) / len(os.listdir(directory))
    print(f'среднее время выполнения оценки одного КТГ - {average_time}')

    #считаем количество совпадений программы с врачом
    number_of_matches = 0
    for res in program_result_dict:
        if program_result_dict[res] == doctor_result_dict[res]:
            number_of_matches += 1
    print(f'совпадений программы с врачом {number_of_matches} из 100')

    #в таблицу сохраняется результат
    #в ней можно будет более подробно рассмотреть общую картину того в каких случаях расхождения между врачом и программой
    wb = load_workbook('ctg.xlsx')
    del wb['Sheet1']
    ws = wb.create_sheet('Sheet1')
    for res in program_result_dict:
        ws.append([res, doctor_result_dict[res], program_result_dict[res]])
    wb.save('ctg.xlsx')
